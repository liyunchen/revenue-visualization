import requests
import json
import time
import openpyxl


###拉伸图
from pyecharts import options as opts
from pyecharts.globals import ThemeType
from pyecharts.charts import Bar

def silder(name,value):
    c = (
        Bar(init_opts=opts.InitOpts(theme=ThemeType.DARK))
            .add_xaxis(xaxis_data=name)
            .add_yaxis("日收入/美元", yaxis_data=value)
            .set_global_opts(
            title_opts=opts.TitleOpts(title="王者荣耀近一个月日收入情况"),
            datazoom_opts=[opts.DataZoomOpts(), opts.DataZoomOpts(type_="inside")],
        )
            .render("王者荣耀近一个月日收入情况.html")
    )

###流水线图
import bar_chart_race as bcr
import pandas as pd
import matplotlib.pyplot as plt
"""
李运辰 2021-3-28

公众号：python爬虫数据分析挖掘

代码不明白的，可以去公众号看文章代码讲解
"""

###时间戳转为字符串
def todate(timeStamp):
    timeStamp = int(timeStamp)
    timeArray = time.localtime(timeStamp)  # 将时间戳转换成元组对象
    time_str = time.strftime('%Y-%m-%d', timeArray)  # 将元组转换成对应的时间格式
    return time_str


###获取app收入排行
def app_revenue_data():
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 6.3; Win64; x64; rv:84.0) Gecko/20100101 Firefox/84.0',
        'cookie':'PHPSESSID=7vg101mb7pggb73562er7o7q8i; qm_check=SxJXQEUSChd2fHd1dRQQeV5EVVwYYhwXenQZd0ZZQFhZU0MQBAMHAhB9WUZdU0QDdAEBEENEbQVmABRIQ28FbwAZEhkVUVhcU18aChIAHAAcABkHHgAbTQ%3D%3D; gr_user_id=52a87de5-03a2-4f57-981d-c1c9df006422; ada35577182650f1_gr_session_id=bb3b8615-b6a8-4c1c-8cd8-ba3741650a44; ada35577182650f1_gr_session_id_bb3b8615-b6a8-4c1c-8cd8-ba3741650a44=true; aso_ucenter=ec4c%2FjtFHU6HuyjWnkUAc0OI%2BzTCFbHADxgOced6RNrLP7V%2BJLCseNT04UwoEtYWBA; USERINFO=ek%2FTQYtglR3rhzkWdCL9Yc1vO0PC82fNO1KdyUJs606aMWgPGnOAOZQYk15sJ74xH%2Fq7n0lcSlqwnXEGAN2zQX9DCdMH5uNtVRBKqGadB9ZvykGL5U7e5WAt4j14GBoAaB4jWSuVBEHkQ5zArXtIAw%3D%3D; ada35577182650f1_gr_last_sent_sid_with_cs1=bb3b8615-b6a8-4c1c-8cd8-ba3741650a44; ada35577182650f1_gr_last_sent_cs1=qm11690520255; AUTHKEY=1TwTdom8ugVdQBxRcDr1ykeMkl9WVJaYFMhQUPVji%2FzErjGGYgPXs4uCcGtTQrA5wwHGbYphwB8N9vZBj9i6V2vxJJyEnmW7pviw9CqjZ99nFZcKR10ILg%3D%3D; syncd=-1398; synct=1616907840.171; ada35577182650f1_gr_cs1=qm11690520255',
       }
    url="https://api.qimai.cn/pred/appMonthPred?analysis=eEcbRhNVVB9RQEB9VwpDUDRGUwVwEwEAAQcIClwODlQEAyETAQ%3D%3D"

    date=['2020-01','2020-02','2020-03','2020-04','2020-05','2020-06','2020-07',
          '2020-08','2020-09','2020-10','2020-11','2020-12','2021-01','2021-02']
    #app名称: 收入情况
    dict={}
    for d in range(0,len(date)):
        data = {
            'device': 'iphone',
            'genre': 36,
            'month': date[d],
        }
        response = requests.post(url, headers=headers, data=data)
        text = json.loads(response.text)
        revenue_data = text['revenue_data']

        for i in revenue_data:
           d_get = dict.get(i['app_name'])

           if d_get== None:#不存在
               #创建
               dict[i['app_name']]=[0]*len(date)

           tem_list = dict[i['app_name']]
           tem_list[d] = i['revenue']
           dict[i['app_name']] = tem_list

    outwb = openpyxl.Workbook()
    outws = outwb.create_sheet(index=0)
    outws.cell(row=1, column=1, value="日期")
    for i in range(0,len(date)):
        outws.cell(row=1, column=i+2, value=date[i])
    ###写入csv
    # 通过遍历keys()来获取所有的键
    count =2
    for k in dict.keys() :
        outws.cell(row=count, column=1, value=k)
        ###写入值
        tem_list = dict[k]
        for j in range(0,len(tem_list)):
            outws.cell(row=count, column=j+2, value=tem_list[j])
        count = count+1
    outwb.save("App收入排行_lyc.xlsx")  # 保存

###绘制流水线图
def pic1():
    ####开始画图
    plt.rcParams['font.sans-serif']=['SimHei'] #显示中文标签
    plt.rcParams['axes.unicode_minus']=False   #解决负号“-”显示为方块的问题

    # 获取数据
    df = pd.read_csv("App收入排行_lyc.csv",index_col=0)
    # 生成动态流水线
    bcr.bar_chart_race(df=df,
                        filename='App收入排行_lyc.mp4', #生成的动态条形图的文件位置
                        orientation='h', #h条形图 v柱状图
                        sort='desc', #降序，asc-升序
                        n_bars=10, #设置最多能显示的条目数
                        fixed_order=False, # 设置固定类目
                        fixed_max=False, #固定数值轴，使其不发生动态变化 True-固定
                        steps_per_period=24, #图像帧数:数值越小，越不流畅,越大，越流畅
                        period_length=20, #设置帧率，单位时间默认为500ms 即为24帧的总时间是500ms
                        end_period_pause=0,#固定值比如年份的停留时间
                        interpolate_period=False,
                        period_label={'x': .80, 'y': .5, 'ha': 'right', 'va': 'center','size':16}, #设置日期标签的时间格式
                        colors='dark12', #设置柱状图颜色颜色，通过在「_colormaps.py」文件中添加颜色信息，即可自定义配置颜色
                        title={'label': 'App收入排行_lyc','size': 18,}, #图表标题
                        bar_size=.95, #条形图高度
                        bar_textposition='inside',#条形图标签文字位置
                        bar_texttemplate='{x:,.0f}', #条形图标签文字格式
                        bar_label_font=16, #条形图标签文字大小
                        tick_label_font=16, #坐标轴标签文字大小
                        tick_template='{x:,.0f}',#坐标轴标签文字格式
                        shared_fontdict={'family': 'Microsoft YaHei','color': 'rebeccapurple'}, #全局字体属性
                        scale='linear',
                        fig=None,
                        writer=None,
                        bar_kwargs={'alpha': .7},#条形图属性，可以设置透明度，边框等
                        fig_kwargs={'figsize': (16, 10), 'dpi': 144},#figsize-设置画布大小，默认(6, 3.5)，dpi-图像分辨率，默认144
                        filter_column_colors=True,#去除条形图重复颜色，True去除,默认为False
                )

###王者荣耀近一个月日收入
def near_month():

    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 6.3; Win64; x64; rv:84.0) Gecko/20100101 Firefox/84.0',
        'cookie': 'PHPSESSID=7vg101mb7pggb73562er7o7q8i; qm_check=SxJXQEUSChd2fHd1dRQQeV5EVVwYYhwXenQZd0ZZQFhZU0MQBAMHAhB9WUZdU0QDdAEBEENEbQVmABRIQ28FbwAZEhkVUVhcU18aChIAHAAcABkHHgAbTQ%3D%3D; gr_user_id=52a87de5-03a2-4f57-981d-c1c9df006422; ada35577182650f1_gr_session_id=bb3b8615-b6a8-4c1c-8cd8-ba3741650a44; ada35577182650f1_gr_session_id_bb3b8615-b6a8-4c1c-8cd8-ba3741650a44=true; aso_ucenter=ec4c%2FjtFHU6HuyjWnkUAc0OI%2BzTCFbHADxgOced6RNrLP7V%2BJLCseNT04UwoEtYWBA; USERINFO=ek%2FTQYtglR3rhzkWdCL9Yc1vO0PC82fNO1KdyUJs606aMWgPGnOAOZQYk15sJ74xH%2Fq7n0lcSlqwnXEGAN2zQX9DCdMH5uNtVRBKqGadB9ZvykGL5U7e5WAt4j14GBoAaB4jWSuVBEHkQ5zArXtIAw%3D%3D; ada35577182650f1_gr_last_sent_sid_with_cs1=bb3b8615-b6a8-4c1c-8cd8-ba3741650a44; ada35577182650f1_gr_last_sent_cs1=qm11690520255; AUTHKEY=1TwTdom8ugVdQBxRcDr1ykeMkl9WVJaYFMhQUPVji%2FzErjGGYgPXs4uCcGtTQrA5wwHGbYphwB8N9vZBj9i6V2vxJJyEnmW7pviw9CqjZ99nFZcKR10ILg%3D%3D; syncd=-1398; synct=1616907840.171; ada35577182650f1_gr_cs1=qm11690520255',
    }
    url = "https://api.qimai.cn/pred/revenue?analysis=dQ51TyxjAEd9WQBJdg5%2BTylecxV9dH1EfVpTDStzU1Z6TCwFflpiWlJXBVl3G0tERARUH0JVRlVeTQF3G1UEB1YJBQQJBgAECyQUCQ%3D%3D&appid=989673964&country=cn&sdate=2021-02-26&edate=2021-03-27"
    response = requests.post(url, headers=headers)
    text = json.loads(response.text)
    data_list = text['data']['list']
    name=[]
    value=[]
    for i in range(len(data_list)-1,-1,-1):

        name.append(todate(str(data_list[i][0])[0:-3]))
        value.append(data_list[i][1])
    ###绘图
    silder(name,value)

###王者荣耀近一年月收入情况
def near_year():
    # df = pd.read_csv("App收入排行_lyc.csv", index_col=0)
    # d = df.tolist
    # print(d)

    f = open("App收入排行_lyc.csv",encoding="utf-8")
    content = f.read()
    rows = content.split('\n')

    name = rows[0].split(",")[1:]
    dict_values = rows[1].split(",")[1:]
    # name = data_top1.index.tolist()
    # dict_values = data_top1.values.tolist()
    # 链式调用
    c = (
        Bar(
            init_opts=opts.InitOpts(  # 初始配置项
                theme=ThemeType.MACARONS,
                animation_opts=opts.AnimationOpts(
                    animation_delay=1000, animation_easing="cubicOut"  # 初始动画延迟和缓动效果
                ))
        )
            .add_xaxis(xaxis_data=name)  # x轴
            .add_yaxis(series_name="王者荣耀近一年月收入情况", yaxis_data=dict_values)  # y轴
            .set_global_opts(
            title_opts=opts.TitleOpts(title='', subtitle='',  # 标题配置和调整位置
                                      title_textstyle_opts=opts.TextStyleOpts(
                                          font_family='SimHei', font_size=25, font_weight='bold', color='red',
                                      ), pos_left="90%", pos_top="10",
                                      ),
            xaxis_opts=opts.AxisOpts(name='月份', axislabel_opts=opts.LabelOpts(rotate=45)),
            # 设置x名称和Label rotate解决标签名字过长使用
            yaxis_opts=opts.AxisOpts(name='月收入/美元'),

        )
            .render("王者荣耀近一年月收入情况.html")
    )


near_year()



